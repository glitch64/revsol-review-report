import pyodbc
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter
from datetime import datetime
import os

# --- Configuration ---
SERVER = 'MSI\\SQLEXPRESS'
DATABASE = 'revsol'
USERNAME = 'sa'
PASSWORD = 'Revs0ls!'
DRIVER = '{ODBC Driver 17 for SQL Server}'  # Update if using a different driver

# --- File Setup ---
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
OUTBOX_DIR = os.path.join(BASE_DIR, 'OUTBOX')
TODAY = datetime.now().strftime('%Y%m%d')
FILENAME = f'RevSolReview_{TODAY}.xlsx'
OUTPUT_PATH = os.path.join(OUTBOX_DIR, FILENAME)

# --- SQL Query ---
QUERY = '''
select 
user_name as [User],
duedate   as [Due Date],
duedatetime as [Due Date Time],
orgname as [Organization],
org_type as [Org Type],
pername [Person],
peremail [Person Email],
subject [Subject],
actnote [Note]
from review
order by 
 user_name,
 duedatetime,
 orgname,
 pername,
type
'''

# Function to clean up Excel-invalid or HTML-formatted cell values
def clean_cell_value(value):
    import html
    import re

    if isinstance(value, str):
        # Decode HTML entities like &nbsp;, &amp;
        value = html.unescape(value)

        # Remove basic HTML tags
        value = re.sub(r'<[^>]+>', '', value)

        # Replace newlines from tags with space (clean rendering)
        value = value.replace('\r', ' ').replace('\n', ' ')

        # Remove ALL control characters except TAB (\x09), LF (\x0A), CR (\x0D)
        value = ''.join(ch for ch in value if ch == '\t' or ch == '\n' or ch == '\r' or ord(ch) >= 32)
    return value

def get_data_from_sql_server():
    conn_str = f'DRIVER={DRIVER};SERVER={SERVER};DATABASE={DATABASE};UID={USERNAME};PWD={PASSWORD}'
    with pyodbc.connect(conn_str) as conn:
        cursor = conn.cursor()
        cursor.execute(QUERY)
        columns = [column[0] for column in cursor.description]
        rows = cursor.fetchall()
        return columns, rows

def create_excel_report(headers, data):
    wb = Workbook()
    ws = wb.active
    ws.title = 'Review Report'
    ws.freeze_panes = "A2"

    # Formatting settings by column name
    column_settings = {
        "User": (20, False),
        "Due Date": (15, False),
        "Due Date Time": (18, False),
        "Organization": (25, False),
        "Org Type": (14, False),
        "Person": (20, False),
        "Person Email": (35, False),
        "Subject": (40, False),
        "Note": (65, True),
    }

    # Write headers with bold font and apply filter
    bold_font = Font(bold=True)
    for col_num, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = bold_font

        # Set column width
        col_letter = get_column_letter(col_num)
        width, wrap = column_settings.get(header, (20, False))
        ws.column_dimensions[col_letter].width = width

        # Apply alignment (top aligned, wrap as specified)
        alignment = Alignment(vertical='top', wrap_text=wrap)
        cell.alignment = alignment

    # Write data rows
    for row_num, row_data in enumerate(data, start=2):
        for col_num, value in enumerate(row_data, start=1):
            cleaned_value = clean_cell_value(value)
            cell = ws.cell(row=row_num, column=col_num, value=cleaned_value)
            header = headers[col_num - 1]
            _, wrap = column_settings.get(header, (20, False))
            cell.alignment = Alignment(vertical='top', wrap_text=wrap)

    # Add auto-filter to header row
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"

    # Ensure OUTBOX directory exists
    os.makedirs(OUTBOX_DIR, exist_ok=True)

    # Set font size 10 globally
    for row in ws.iter_rows():
        for cell in row:
            if cell.value is not None:
                current_font = cell.font or Font()
                cell.font = Font(name=current_font.name, size=10, bold=current_font.bold)

    wb.save(OUTPUT_PATH)
    print(f'\n✅ Excel report saved to: {OUTPUT_PATH}')

def main():
    try:
        headers, data = get_data_from_sql_server()
        create_excel_report(headers, data)
    except Exception as e:
        print(f'\n❌ Error: {e}')

if __name__ == '__main__':
    main()

